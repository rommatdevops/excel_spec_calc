import openpyxl
from openpyxl.styles import Font, Color, colors, fills

FIRST_LINE = 6
LAST_LINE = 8984

FILENAME = "file"

def openFile():
    wb = openpyxl.load_workbook(FILENAME + '.xlsx')
    return wb

def appendAllValues(column_number):
    vhodjenia = {}
    row_and_value = {}
    list = []
    for i in range(FIRST_LINE, LAST_LINE, 1):
        row_and_value[i] = sheet.cell(row=i, column=column_number).value
        if sheet.cell(row=i, column=column_number).value in list:
            vhodjenia[sheet.cell(row=i, column=column_number).value] +=1

        if sheet.cell(row=i, column=column_number).value not in list:
            vhodjenia[sheet.cell(row=i, column=column_number).value] = 1

        list.append(sheet.cell(row=i, column=column_number).value)
    return vhodjenia, row_and_value

def sortValues(pryhod, rozhod, rows_pryhod, rows_rozhod):
    list_items = []
    for el in pryhod ^ rozhod:
        if el[0] == str(' '):
            continue
        if el[0] in list_items:
            continue
        list_items.append(el[0])

    printPreResultText(list_items)

    for el in list_items:
        if el == str(' '):
            continue
        print('------------------------------------------------------------------------------------')
        print(el)
        line_pryhod = showLine(el, rows_pryhod, 5)
        line_rozhod = showLine(el, rows_rozhod, 6)
        print(line_pryhod)
        print(line_rozhod)
        for el1 in line_pryhod:
            fillRow(el1, 5)
        for el1 in line_rozhod:
            fillRow(el1, 6)

def printPreResultText(list_items):
    print("Resume")
    print("В наступних сумах є неспівпадіння:")
    print(list_items)
    write_log(list_items)
    print("Далі по кожному значенню детальніше, із зазначенням рядка, в якому вони знаходяться:")

def write_log(text):
    f = open("log.txt", "a")
    f.write("В наступних сумах є неспівпадіння: " + str(text) + "\n")
    f.close()

def showLine(value_search, list, column):
    keys = [k for k, v in list.items() if v == float(value_search)]
    return keys

def fillRow(row, column):
    redFill = fills.PatternFill(patternType='solid', fgColor=colors.Color(rgb='00FF0000'))
    sheet.cell(row=row, column=column).fill = redFill

if __name__ == '__main__':
    file = openFile()
    sheet = file['Лист 1']
    rozhod = sheet['F6']

    pryhod_dict = set(appendAllValues(5)[0].items())
    rozhod_dict= set(appendAllValues(6)[0].items())

    print(pryhod_dict ^ rozhod_dict)

    rows_pryhod = appendAllValues(5)[1]
    rows_rozhod = appendAllValues(6)[1]

    sortValues(pryhod_dict, rozhod_dict, rows_pryhod, rows_rozhod)
    file.save(FILENAME + "_checked.xlsx")


